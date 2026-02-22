"""Controlador para exportación de liquidaciones a Excel y PDF."""
import os
from datetime import datetime
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.pdfgen import canvas
from sqlalchemy.orm import joinedload
from app.models.payroll_summary import PayrollSummary
from app.models.payroll_detail import PayrollDetail
from app.models.app_user import AppUser
from app.models.driver import Driver
from app.models.base import db


class PayrollExportController:
    """Controlador para exportación de liquidaciones."""
    
    EXPORT_DIR = 'exports/payroll'
    
    @staticmethod
    def _ensure_export_dir():
        """Asegurar que el directorio de exportación existe."""
        if not os.path.exists(PayrollExportController.EXPORT_DIR):
            os.makedirs(PayrollExportController.EXPORT_DIR)
    
    @staticmethod
    def export_to_excel(summary_id):
        """Exportar resumen de liquidación a Excel."""
        from app.controllers.payroll_calculation import PayrollCalculationController
        PayrollExportController._ensure_export_dir()
        
        # Obtener datos con eager loading de las relaciones
        summary = PayrollSummary.query.options(
            joinedload(PayrollSummary.driver).joinedload(Driver.user),
            joinedload(PayrollSummary.period)
        ).get(summary_id)
        
        if not summary:
            raise ValueError(f"No se encontró el resumen con ID {summary_id}")
        
        _, details = PayrollCalculationController.get_summary_details(summary_id)
        
        # Obtener nombre completo del driver
        driver_full_name = f"{summary.driver.user.name} {summary.driver.user.surname}"
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Liquidación {driver_full_name}"
        
        # Estilos
        header_font = Font(bold=True, size=14)
        subheader_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = 'LIQUIDACIÓN DE SUELDO'
        ws['A1'].font = header_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:D1')
        
        # Información del período
        row = 3
        ws[f'A{row}'] = 'Período:'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = f"{summary.period.year}-{summary.period.month:02d}"
        
        row += 1
        ws[f'A{row}'] = 'Chofer:'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = driver_full_name
        
        row += 1
        ws[f'A{row}'] = 'CUIL:'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = summary.driver.cuil or 'N/A'
        
        row += 1
        ws[f'A{row}'] = 'Comisión aplicada:'
        ws[f'A{row}'].font = bold_font
        ws[f'B{row}'] = f"{summary.driver_commission_percentage}%"
        
        # Detalle de cálculo
        row += 2
        ws[f'A{row}'] = 'DETALLE DEL CÁLCULO'
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        row += 1
        headers = ['Concepto', 'Descripción', 'Monto', 'Tipo']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Agrupar detalles por tipo
        trip_details = []
        expense_reimburse = []
        expense_deduct = []
        advances = []
        other_items = []
        
        for detail in details:
            if detail.detail_type == 'trip_commission':
                trip_details.append(detail)
            elif detail.detail_type == 'expense_reimburse':
                expense_reimburse.append(detail)
            elif detail.detail_type == 'expense_deduct':
                expense_deduct.append(detail)
            elif detail.detail_type in ['advance', 'client_advance']:
                advances.append(detail)
            elif detail.detail_type in [
                'adjustment',
                'other_item',
                'other_item_adjustment',
                'other_item_bonus',
                'other_item_charge',
                'other_item_fine'
            ]:
                other_items.append(detail)
        
        # Escribir detalles de viajes
        if trip_details:
            row += 1
            ws[f'A{row}'] = 'COMISIÓN POR VIAJES'
            ws[f'A{row}'].font = bold_font
            ws.merge_cells(f'A{row}:D{row}')
            
            for detail in trip_details:
                row += 1
                ws[f'A{row}'] = 'Viaje'
                ws[f'B{row}'] = detail.description
                ws[f'C{row}'] = float(detail.amount)
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = '+'
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Escribir gastos a reintegrar
        if expense_reimburse:
            row += 1
            ws[f'A{row}'] = 'GASTOS A REINTEGRAR'
            ws[f'A{row}'].font = bold_font
            ws.merge_cells(f'A{row}:D{row}')
            
            for detail in expense_reimburse:
                row += 1
                ws[f'A{row}'] = 'Reintegro'
                ws[f'B{row}'] = detail.description
                ws[f'C{row}'] = float(detail.amount)
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = '+'
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Escribir gastos a descontar
        if expense_deduct:
            row += 1
            ws[f'A{row}'] = 'GASTOS A DESCONTAR'
            ws[f'A{row}'].font = bold_font
            ws.merge_cells(f'A{row}:D{row}')
            
            for detail in expense_deduct:
                row += 1
                ws[f'A{row}'] = 'Descuento'
                ws[f'B{row}'] = detail.description
                ws[f'C{row}'] = float(detail.amount)
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = '-'
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Escribir adelantos
        if advances:
            row += 1
            ws[f'A{row}'] = 'ADELANTOS'
            ws[f'A{row}'].font = bold_font
            ws.merge_cells(f'A{row}:D{row}')
            
            for detail in advances:
                row += 1
                ws[f'A{row}'] = 'Adelanto'
                ws[f'B{row}'] = detail.description
                ws[f'C{row}'] = float(detail.amount)
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = '-'
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Escribir otros conceptos
        if other_items:
            row += 1
            ws[f'A{row}'] = 'OTROS CONCEPTOS'
            ws[f'A{row}'].font = bold_font
            ws.merge_cells(f'A{row}:D{row}')
            
            for detail in other_items:
                row += 1
                ws[f'A{row}'] = 'Concepto'
                ws[f'B{row}'] = detail.description
                ws[f'C{row}'] = float(detail.amount)
                ws[f'C{row}'].number_format = '$#,##0.00'
                ws[f'D{row}'] = '+' if detail.amount >= 0 else '-'
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = thin_border
        
        # Resumen de totales
        row += 2
        ws[f'A{row}'] = 'RESUMEN'
        ws[f'A{row}'].font = subheader_font
        ws[f'A{row}'].fill = subheader_fill
        ws.merge_cells(f'A{row}:D{row}')
        
        totals = [
            ('Comisión por viajes', summary.commission_from_trips),
            ('Gastos a reintegrar', summary.expenses_to_reimburse),
            ('Gastos a descontar', -summary.expenses_to_deduct),
            ('Mínimo garantizado', summary.guaranteed_minimum_applied),
            ('Adelantos', -summary.advances_deducted),
            ('Otros conceptos', summary.other_items_total),
        ]
        
        for label, amount in totals:
            row += 1
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = bold_font
            ws[f'C{row}'] = float(amount)
            ws[f'C{row}'].number_format = '$#,##0.00'
            for col in range(1, 4):
                ws.cell(row=row, column=col).border = thin_border
        
        # Saldo a favor y saldo en contra
        row += 2
        ws[f'A{row}'] = 'Saldo a favor'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'C{row}'] = float(summary.balance_in_favor)
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].font = Font(bold=True, size=12)
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
        ws[f'A{row}'] = 'Saldo en contra'
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'C{row}'] = float(summary.balance_against)
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].font = Font(bold=True, size=12)
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
        
        # Total final
        row += 1
        ws[f'A{row}'] = 'TOTAL A PAGAR'
        ws[f'A{row}'].font = Font(bold=True, size=14)
        ws[f'A{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        ws[f'C{row}'] = float(summary.total_amount)
        ws[f'C{row}'].number_format = '$#,##0.00'
        ws[f'C{row}'].font = Font(bold=True, size=14)
        ws[f'C{row}'].fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = thin_border
        
        # Ajustar anchos de columna
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 50
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
        
        # Guardar archivo
        # Formato: Liquidacion_EstadoResumen_NombreChofer_Periodo.xlsx
        driver_name = f"{summary.driver.user.name}_{summary.driver.user.surname}"
        driver_name = driver_name.replace(' ', '_')
        period_name = f"{summary.period.month:02d}-{summary.period.year}"
        status_map = {
            'draft': 'Borrador',
            'pending_approval': 'PendienteAprobacion',
            'approved': 'Aprobado',
            'error': 'Error',
            'calculation_pending': 'CalculoPendiente'
        }
        status_name = status_map.get(summary.status, summary.status)
        filename = f"Liquidacion_{status_name}_{driver_name}_{period_name}.xlsx"
        filepath = os.path.join(PayrollExportController.EXPORT_DIR, filename)
        
        wb.save(filepath)
        
        # Actualizar resumen
        summary.export_format = 'excel'
        summary.export_path = filepath
        db.session.commit()
        
        return filepath
    
    @staticmethod
    def export_to_pdf(summary_id):
        """Exportar resumen de liquidación a PDF."""
        from app.controllers.payroll_calculation import PayrollCalculationController
        PayrollExportController._ensure_export_dir()
        
        # Obtener datos con eager loading de las relaciones
        summary = PayrollSummary.query.options(
            joinedload(PayrollSummary.driver).joinedload(Driver.user),
            joinedload(PayrollSummary.period)
        ).get(summary_id)
        
        if not summary:
            raise ValueError(f"No se encontró el resumen con ID {summary_id}")
        
        _, details = PayrollCalculationController.get_summary_details(summary_id)
        
        # Obtener nombre completo del driver
        driver_full_name = f"{summary.driver.user.name} {summary.driver.user.surname}"
        
        # Validar que esté aprobado
        if summary.status != 'approved':
            raise ValueError("Solo se pueden exportar a PDF resúmenes aprobados")
        
        # Crear PDF
        # Formato: Resumen_NombreChofer_Periodo.pdf
        driver_name = f"{summary.driver.user.name}_{summary.driver.user.surname}"
        driver_name = driver_name.replace(' ', '_')
        period_name = f"{summary.period.month:02d}-{summary.period.year}"
        filename = f"Resumen_{driver_name}_{period_name}.pdf"
        filepath = os.path.join(PayrollExportController.EXPORT_DIR, filename)
        
        doc = SimpleDocTemplate(filepath, pagesize=letter)
        story = []
        styles = getSampleStyleSheet()
        
        # Título
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#366092'),
            spaceAfter=30,
            alignment=1  # Center
        )
        story.append(Paragraph('LIQUIDACIÓN DE SUELDO', title_style))
        
        # Información del período
        info_data = [
            ['Período:', f"{summary.period.year}-{summary.period.month:02d}"],
            ['Chofer:', driver_full_name],
            ['CUIL:', summary.driver.cuil or 'N/A'],
            ['CBU:', summary.driver.cbu or 'N/A'],
            ['Comisión:', f"{summary.driver_commission_percentage}%"],
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Detalle
        story.append(Paragraph('DETALLE DEL CÁLCULO', styles['Heading2']))
        story.append(Spacer(1, 12))
        
        detail_data = [['Concepto', 'Descripción', 'Monto']]
        
        for detail in details:
            sign = '+' if detail.amount >= 0 else '-'
            detail_data.append([
                detail.detail_type.replace('_', ' ').title(),
                detail.description[:50],
                f"${abs(float(detail.amount)):,.2f} {sign}"
            ])
        
        detail_table = Table(detail_data, colWidths=[1.5*inch, 3.5*inch, 1.5*inch])
        detail_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))
        story.append(detail_table)
        story.append(Spacer(1, 20))
        
        # Resumen
        story.append(Paragraph('RESUMEN', styles['Heading2']))
        story.append(Spacer(1, 12))
        
        summary_data = [
            ['Comisión por viajes', f"${float(summary.commission_from_trips):,.2f}"],
            ['Gastos a reintegrar', f"${float(summary.expenses_to_reimburse):,.2f}"],
            ['Gastos a descontar', f"-${float(summary.expenses_to_deduct):,.2f}"],
            ['Mínimo garantizado', f"${float(summary.guaranteed_minimum_applied):,.2f}"],
            ['Adelantos', f"-${float(summary.advances_deducted):,.2f}"],
            ['Otros conceptos', f"${float(summary.other_items_total):,.2f}"],
            ['', ''],
            ['Saldo a favor', f"${float(summary.balance_in_favor):,.2f}"],
            ['Saldo en contra', f"${float(summary.balance_against):,.2f}"],
            ['', ''],
            ['TOTAL A PAGAR', f"${float(summary.total_amount):,.2f}"]
        ]
        
        summary_table = Table(summary_data, colWidths=[4*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -4), colors.lightgrey),
            ('BACKGROUND', (0, -3), (0, -2), colors.HexColor('#D9EAD3')),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#FFC000')),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTNAME', (0, -3), (0, -2), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -3), (-1, -2), 12),
            ('FONTSIZE', (0, -1), (-1, -1), 14),
            ('GRID', (0, 0), (-1, -4), 1, colors.black),
            ('GRID', (0, -3), (-1, -2), 1, colors.black),
            ('GRID', (0, -1), (-1, -1), 2, colors.black),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12)
        ]))
        story.append(summary_table)
        
        # Pie de página
        story.append(Spacer(1, 40))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            textColor=colors.grey,
            alignment=1
        )
        story.append(Paragraph(
            f"Documento generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}",
            footer_style
        ))
        
        # Generar PDF
        doc.build(story)
        
        # Actualizar resumen
        summary.export_format = 'pdf'
        summary.export_path = filepath
        from app.models.base import db
        db.session.commit()
        
        return filepath
